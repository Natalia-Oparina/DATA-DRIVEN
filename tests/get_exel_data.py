from openpyxl import load_workbook

exel_path = r'C:\DATA-DRIVEN\data\data.xlsx'

wb = load_workbook(exel_path)

ws = wb.worksheets[0]

login_form_parameters = []

for x in range(2, ws.max_row + 1):
    username = ws.cell(column=1, row=x).value
    password = ws.cell(column=2, row=x).value
    checkpoint = ws.cell(column=3, row=x).value

    tuple = username, password, checkpoint
    login_form_parameters.append(tuple)
    